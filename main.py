from tkinter import *
from assets.bg_loader import load_bg_image
from assets.bg_base64 import background_image
from assets.colors import *
from tkinter import messagebox
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl import load_workbook
import os

window = Tk()
window.title("Student Score Tracker by Khian Victory D. Calderon")
window.state("zoomed")

# Load background
load_bg_image(window, background_image)

# Create the Listbox widget
my_list = Listbox(window, yscrollcommand=None, bg=color_darker_blue, fg="white", font=("Verdana", 20))

# Functions
student_count = 0
score_summation = 0

def save_to_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Scores"

    ws.append(["Name", "Score", "Remark"])
    header_font = Font(bold=True)
    for col in range(1, 4):  # A to C
        ws.cell(row=1, column=col).font = header_font

    total_score = 0
    student_count = 0

    for i in range(my_list.size()):
        item = my_list.get(i)

        if "—" in item or item.startswith("Average:"):
            continue

        parts = item.split(" : ")
        if len(parts) == 3:
            name = parts[0].strip()
            score = parts[1].strip().replace(" / 100", "")
            remark = parts[2].strip()
            ws.append([name, score, remark])
            total_score += int(score)
            student_count += 1

    ws.append([])  # Empty row before average
    if student_count > 0:
        avg_row = ws.max_row + 1
        avg_cell = ws.cell(row=avg_row + 1, column=1, value="Average")
        avg_value_cell = ws.cell(row=avg_row + 1, column=2, value=round(total_score / student_count, 2))

        avg_cell.font = Font(bold=True)
        avg_value_cell.font = Font(bold=True)
        avg_value_cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 20

    try:
        wb.save("student_scores.xlsx")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to save file: {e}")

def load_from_excel():
    if not os.path.exists("student_scores.xlsx"):
        return  # No file to load

    try:
        wb = load_workbook("student_scores.xlsx")
        ws = wb.active

        my_list.delete(0, END)  # Clear current listbox
        total_score = 0
        student_count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row == (None, None, None):  # Skip empty row before average
                continue
            if row[0] == "Average":
                break  # Stop at average row

            name, score, remark = row
            my_list.insert(END, f"{name} : {score} / 100 : {remark}")
            total_score += int(score)
            student_count += 1

        if student_count > 0:
            my_list.insert(END, f"Average: {total_score / student_count:.2f}")

    except Exception as e:
        messagebox.showerror("Error", f"Failed to load data from Excel: {e}")

def handle_add_student():
    name = student_name.get().strip()
    student_score = score.get().strip()
    global student_count
    global score_summation

    if not name or not student_score:
        messagebox.showerror("Error adding student", "Both fields are required.")
        return

    if not re.fullmatch(r"[A-Za-z .]+", name):
        messagebox.showerror("Error adding student", "Name can only contain letters, spaces, and periods.")
        return
    
    if not student_score.isdigit():
        messagebox.showerror("Error adding student", "Score must be a number.")
        return

    # Duplicate check
    for i in my_list.get(0, END):
        if i.startswith(name + "."):
            messagebox.showerror("Error adding student", "Name already in the list.")
            return

    # Add to list
    score_val = int(student_score)
    remark = (
        "Wow Perfect!" if score_val == 100 else
        "Amazing!" if 95 <= score_val <= 99 else
        "Pass" if 75 <= score_val <= 94 else
        "Failed" if 20 <= score_val <= 74 else
        "Yohh what happened?" if 1 <= score_val <= 19 else
        "Bruhh, what?" if score_val == 0 else
        "Invalid Score"
    )
    
    # Update stats
    score_summation += score_val
    student_count += 1

    # Remove previous average + separator if they exist
    if my_list.size() >= 2:
        my_list.delete(END)

    # Insert new student
    my_list.insert(END, f"{name} : {student_score} / 100 : {remark}")
    my_list.insert(END, f"Average: {score_summation / student_count:.2f}")

    # Clear input fields
    student_name.delete(0, END)
    score.delete(0, END)
    
    # Saves to excel
    save_to_excel()

# Initially loads the excel
load_from_excel()

# Title label
Label(
    window, 
    text="Student Score Tracker", 
    font=("Verdana", 20),
    padx=10, 
    pady=10, 
    bg=color_blue, 
    fg="white"
).pack(pady=20)

container = Frame(window, padx=10, pady=10, bg=color_blue)
container.pack(pady=20, padx=80)

Label(container, text="Student Name: ", bg=color_blue, fg="white", font=("Verdana", 20)).grid(row=0, column=0)
student_name = Entry(container, bg=color_light_blue, fg="white", font=("Verdana", 20))
student_name.grid(row=0, column=1, padx=(10, 0), pady = 5)

Label(container, text="Score: ", bg=color_blue, fg="white", font=("Verdana", 20)).grid(row=1, column=0)
score = Entry(container, bg=color_light_blue, fg="white", font=("Verdana", 20))
score.grid(row=1, column=1, padx=(10, 0), pady = 5)

Button(container, text="Add Student", bg=color_light_blue, activebackground=color_ligter_blue, fg="white", font=("Verdana", 20), command=handle_add_student).grid(row=2, column=0, columnspan=2, padx=(10, 0), pady = 10)

# ADD SCROLLBAR
scrollBar = Scrollbar(window)
scrollBar.pack(side=RIGHT, fill=Y, padx = (0, 20), pady = 20)
my_list.pack(fill = BOTH, expand=True, padx = (20, 0), pady = 20)
scrollBar.config(command=my_list.yview)

window.mainloop()
